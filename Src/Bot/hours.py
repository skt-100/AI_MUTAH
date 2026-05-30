import json
from openpyxl import load_workbook
import settings as st
from data_normalize import clean_all

workbook = load_workbook(st.HOURS_EXCEL, read_only=True)
sheet = workbook.active

hours_list = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        name          = clean_all(row[0])
        total         = row[1]  if row[1]  else "غير محدد"
        competency    = row[2]  if row[2]  else "غير محدد"
        remedial      = row[3]  if row[3]  else "غير محدد"
        univ_mand     = row[4]  if row[4]  else "غير محدد"
        univ_hum      = row[5]  if row[5]  else "غير محدد"
        univ_soc      = row[6]  if row[6]  else "غير محدد"
        univ_tech     = row[7]  if row[7]  else "غير محدد"
        col_mand      = row[8]  if row[8]  else "غير محدد"
        spec_mand     = row[9]  if row[9]  else "غير محدد"
        spec_elec     = row[10] if row[10] else "غير محدد"

        # ── النص الطبيعي: اسم التخصص في البداية + جمل واضحة ─────────────
        parts = [f"خطة تخصص {name} الدراسية:"]
        parts.append(f"إجمالي ساعات التخصص {total} ساعة.")

        if spec_mand != "غير محدد":
            parts.append(f"ساعات تخصص إجباري {spec_mand} ساعة.")
        if spec_elec != "غير محدد":
            parts.append(f"ساعات تخصص اختياري {spec_elec} ساعات.")
        if col_mand != "غير محدد":
            parts.append(f"ساعات كلية إجباري {col_mand} ساعة.")
        if univ_mand != "غير محدد":
            parts.append(f"ساعات جامعة إجباري {univ_mand} ساعة.")
        if univ_hum != "غير محدد":
            parts.append(f"ساعات جامعة اختياري علوم إنسانية {univ_hum} ساعات.")
        if univ_soc != "غير محدد":
            parts.append(f"ساعات جامعة اختياري علوم اجتماعية {univ_soc} ساعات.")
        if univ_tech != "غير محدد":
            parts.append(f"ساعات جامعة اختياري علوم تكنولوجيا {univ_tech} ساعات.")
        if competency != "غير محدد":
            parts.append(f"ساعات امتحان الكفاءة {competency}.")
        if remedial != "غير محدد":
            parts.append(f"ساعات الاستدراكي {remedial}.")

        # تكرار اسم التخصص في النهاية يجعل الـ embedding يرتكز عليه كمعرّف أساسي
        parts.append(f"هذه المعلومات تخص تخصص {name} فقط.")
        text = " ".join(parts)

        hours_list.append({
            "النوع":     "ساعات_تخصص",
            "التخصص":   name,
            "النص":      text,
        })

with open(st.HOURS_JSON, "w", encoding="utf-8") as f:
    json.dump(hours_list, f, ensure_ascii=False, indent=2)

print(f"تم! {len(hours_list)} تخصص")
for h in hours_list:
    print(f"  • {h['النص']}")
