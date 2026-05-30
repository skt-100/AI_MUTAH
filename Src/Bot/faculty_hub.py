import json
from openpyxl import load_workbook
import settings as st
from data_normalize import clean_all

workbook = load_workbook(st.FACULTY_EXCEL, read_only=True)
sheet = workbook.active

faculty_list = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[0]:
        name       = clean_all(row[0], is_name=True)
        rank       = clean_all(row[1]) if row[1] else "غير محدد"
        department = clean_all(row[2]) if row[2] else "غير محدد"
        email      = row[3].strip() if row[3] else "غير محدد"

        # ── النص الطبيعي: يشبه كيف يسأل الطالب أو يبحث ──────────────────
        text = (
            f"{name}، {rank} في قسم {department}. "
            f"البريد الإلكتروني: {email}"
        )

        faculty_list.append({
            "النوع":  "عضو_هيئة_تدريس",
            "الاسم":  name,
            "الرتبة": rank,
            "القسم":  department,
            "البريد": email,
            "النص":   text,
        })

with open(st.FACULTY_JSON, "w", encoding="utf-8") as f:
    json.dump(faculty_list, f, ensure_ascii=False, indent=2)

print(f"تم! {len(faculty_list)} عضو هيئة تدريس")
print("مثال:", faculty_list[0]["النص"])
