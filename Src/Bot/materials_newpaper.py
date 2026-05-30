import json
from openpyxl import load_workbook
import settings as st
from data_normalize import clean_all

workbook = load_workbook(st.MATERIALS_EXCEL, read_only=True)
sheet    = workbook.active

# ── المرحلة 1: تجميع الشعب لكل مادة ────────────────────────────────────
materials_dict = {}

for row in sheet.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue

    mat_num    = str(row[0]).strip()
    mat_name   = clean_all(str(row[1]))          if row[1]  else "غير محدد"
    credit_hrs = str(row[2]).strip()             if row[2]  else "غير محدد"
    prereq     = str(row[3]).strip()             if row[3]  else "لا يوجد"
    spec       = clean_all(str(row[4]))          if row[4]  else "غير محدد"
    instructor = clean_all(str(row[5]), is_name=True) if row[5] else "غير محدد"
    sec_num    = str(row[6]).strip()             if row[6]  else "غير محدد"
    capacity   = str(row[7]).strip()             if row[7]  else "غير محدد"
    students   = str(row[8]).strip()             if row[8]  else "غير محدد"
    days       = clean_all(str(row[9]))          if row[9]  else "غير محدد"
    time_slot  = str(row[10]).strip()            if row[10] else "غير محدد"
    location   = clean_all(str(row[11]))         if row[11] else "غير محدد"
    mode       = clean_all(str(row[12]))         if row[12] else "غير محدد"
    status     = clean_all(str(row[13]))         if row[13] else "غير محدد"

    section = {
        "رقم_الشعبة":  sec_num,
        "مدرس_المادة": instructor,
        "سعة_الشعبة":  capacity,
        "عدد_الطلاب":  students,
        "الايام":       days,
        "الساعة":       time_slot,
        "الموقع":       location,
        "نمط_التدريس": mode,
        "حالة_الشعبة": status,
    }

    if mat_num not in materials_dict:
        materials_dict[mat_num] = {
            "رقم_المادة":     mat_num,
            "اسم_المادة":     mat_name,
            "عدد_الساعات":    credit_hrs,
            "المتطلب_السابق": prereq,
            "التخصص":         spec,
            "الشعب":          [],
        }
    materials_dict[mat_num]["الشعب"].append(section)


# ── المرحلة 2: chunk واحد لكل مادة يحتوي على كل شعبها ──────────────────
materials_list = []

for mat in materials_dict.values():
    mat_num    = mat["رقم_المادة"]
    mat_name   = mat["اسم_المادة"]
    credit_hrs = mat["عدد_الساعات"]
    prereq     = mat["المتطلب_السابق"]
    spec       = mat["التخصص"]
    sections   = mat["الشعب"]
    total      = len(sections)

    # ── رأس المادة ──────────────────────────────────────────────────────
    header = (
        f"مادة { mat_name } | "
        f"رقم المادة : { mat_num } | "
        f"التخصص : { spec } | "
        f"الساعات المعتمدة : { credit_hrs } | "
        f"المتطلب السابق : { prereq } | "
        f"عدد الشعب : { total }"
    )

    # ── تفاصيل كل شعبة ──────────────────────────────────────────────────
    sections_lines = []
    for sec in sections:
        location_part = (
            f"قاعة { sec['الموقع'] }"
            if sec["الموقع"] not in ("غير محدد", "")
            else "الموقع غير محدد"
        )

        line = (
            f"  الشعبة { sec['رقم_الشعبة'] } : "
            f"المدرس : { sec['مدرس_المادة'] } | "
            f"الايام : { sec['الايام'] } | "
            f"الوقت : { sec['الساعة'] } | "
            f"{ location_part } | "
            f"نمط التدريس : { sec['نمط_التدريس'] } | "
            f"حالة الشعبة : { sec['حالة_الشعبة'] } | "
            f"السعة : { sec['سعة_الشعبة'] } طالب | "
            f"المسجلون : { sec['عدد_الطلاب'] } طالب"
        )
        sections_lines.append(line)

    # ── النص الكامل = رأس + كل الشعب ────────────────────────────────────
    text = header + "\n" + "\n".join(sections_lines)

    materials_list.append({
        "النوع":          "مادة_كاملة",
        "رقم_المادة":    mat_num,
        "اسم_المادة":    mat_name,
        "التخصص":        spec,
        "إجمالي_الشعب": total,
        "النص":          text,
    })

with open(st.MATERIALS_JSON, "w", encoding="utf-8") as f:
    json.dump(materials_list, f, ensure_ascii=False, indent=2)

print(f"تم! { len(materials_list) } مادة")
print("\nمثال على النص:")
print(materials_list[0]["النص"])
